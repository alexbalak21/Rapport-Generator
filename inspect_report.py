import os
import openpyxl
from docx import Document
from app.core.excel_reader import ExcelReader
from app.core.report_generator import ReportGenerator

print('cwd', os.getcwd())

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active
print('headers', [cell.value for cell in ws[1]])
print('row4', [cell.value for cell in ws[4]])

reader = ExcelReader('data.xlsx')
reader.load()
print('reader columns', reader.get_columns())
print('reader row4', reader.get_row_as_dict(4))

if os.path.exists('template.docx'):
    output = 'tmp_report.docx'
    generator = ReportGenerator('data.xlsx', 'template.docx', 'app/mappings/data.json')
    generator.generate(4, output)
    print('generated', output)
    doc = Document(output)
    for i, p in enumerate(doc.paragraphs):
        if p.text.strip():
            print(i, repr(p.text))
